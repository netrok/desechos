# inventario/api/viewsets.py
from io import BytesIO

from django.http import HttpResponse
from django.utils.timezone import localdate
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import AllowAny, IsAuthenticatedOrReadOnly
from rest_framework.response import Response

from inventario.models import Categoria, InventarioItem, MotivoBaja, Ubicacion
from .serializers import (
    CategoriaSerializer,
    InventarioItemSerializer,
    MotivoBajaSerializer,
    UbicacionSerializer,
)


# --- Permiso “staff” simple (sin meternos a grupos todavía) ---
class IsStaffUser(IsAuthenticatedOrReadOnly):
    """
    - Lectura: pública
    - Escritura: autenticado (por IsAuthenticatedOrReadOnly)
    - Acciones sensibles: se controlan por permission_classes en @action
    """
    pass


def _is_staff(user) -> bool:
    return bool(user and user.is_authenticated and user.is_staff)


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all().order_by("nombre")
    serializer_class = CategoriaSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["nombre"]
    ordering_fields = ["nombre"]


class UbicacionViewSet(viewsets.ModelViewSet):
    queryset = Ubicacion.objects.all().order_by("nombre")
    serializer_class = UbicacionSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["nombre"]
    ordering_fields = ["nombre"]


class MotivoBajaViewSet(viewsets.ModelViewSet):
    queryset = MotivoBaja.objects.all().order_by("nombre")
    serializer_class = MotivoBajaSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["nombre"]
    ordering_fields = ["nombre"]


class InventarioItemViewSet(viewsets.ModelViewSet):
    serializer_class = InventarioItemSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["categoria", "ubicacion", "estado", "activo", "motivo_baja"]
    search_fields = [
        "codigo",
        "serie",
        "modelo",
        "etiqueta_interna",
        "marca",
        "responsable",
        "observaciones",
    ]
    ordering_fields = [
        "codigo",
        "fecha_alta",
        "fecha_baja",
        "marca",
        "modelo",
        "precio_sugerido_venta",
    ]

    def get_queryset(self):
        return (
            InventarioItem.objects.select_related("categoria", "ubicacion", "motivo_baja")
            .all()
            .order_by("-fecha_alta", "codigo")
        )

    # ----------------------------
    # Export XLSX (SOLO STAFF)
    # ----------------------------
    @action(detail=False, methods=["get"], url_path="export/xlsx")
    def export_xlsx(self, request):
        if not _is_staff(request.user):
            return Response({"detail": "Solo staff."}, status=403)

        qs = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            "Código",
            "Categoría",
            "Ubicación",
            "Estado",
            "Marca",
            "Modelo",
            "Serie",
            "Etiqueta interna",
            "Responsable",
            "Activo",
            "Precio sugerido venta",
            "Fecha alta",
            "Fecha baja",
            "Motivo baja",
        ]
        ws.append(headers)

        for it in qs:
            ws.append(
                [
                    it.codigo,
                    it.categoria.nombre if it.categoria_id else "",
                    it.ubicacion.nombre if it.ubicacion_id else "",
                    it.get_estado_display(),  # ✅ bonito
                    it.marca or "",
                    it.modelo or "",
                    it.serie or "",
                    it.etiqueta_interna or "",
                    it.responsable or "",
                    "Sí" if it.activo else "No",
                    float(it.precio_sugerido_venta) if it.precio_sugerido_venta is not None else None,
                    it.fecha_alta.isoformat() if it.fecha_alta else "",
                    it.fecha_baja.isoformat() if it.fecha_baja else "",
                    it.motivo_baja.nombre if it.motivo_baja_id else "",
                ]
            )

        # ancho decente
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        filename = f"inventario_{localdate().isoformat()}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # ----------------------------
    # Export PDF listado (SOLO STAFF)
    # ----------------------------
    @action(detail=False, methods=["get"], url_path="export/pdf")
    def export_pdf(self, request):
        if not _is_staff(request.user):
            return Response({"detail": "Solo staff."}, status=403)

        qs = self.filter_queryset(self.get_queryset())

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
            title="Inventario de desechos",
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph("Inventario de desechos electrónicos", styles["Title"]),
            Spacer(1, 8),
        ]

        headers = [
            "Foto",
            "Código",
            "Categoría",
            "Ubicación",
            "Estado",
            "Marca",
            "Modelo",
            "Serie",
            "Activo",
            "Precio sugerido",
        ]
        data = [headers]

        for it in qs:
            img_cell = ""
            if it.foto and hasattr(it.foto, "path"):
                try:
                    img_cell = RLImage(it.foto.path, width=40, height=40)
                except Exception:
                    img_cell = ""

            data.append(
                [
                    img_cell,
                    it.codigo or "",
                    it.categoria.nombre if it.categoria_id else "",
                    it.ubicacion.nombre if it.ubicacion_id else "",
                    it.get_estado_display(),  # ✅ bonito
                    it.marca or "",
                    it.modelo or "",
                    it.serie or "",
                    "Sí" if it.activo else "No",
                    f"${it.precio_sugerido_venta:.2f}" if it.precio_sugerido_venta is not None else "",
                ]
            )

        col_widths = [50, 70, 90, 90, 70, 70, 90, 90, 55, 85]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ]
            )
        )

        story.append(table)
        doc.build(story)

        pdf = buffer.getvalue()
        buffer.close()

        filename = f"inventario_{localdate().isoformat()}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ----------------------------
    # Ficha PDF por item (SOLO STAFF)
    # /api/items/{id}/ficha/pdf/
    # ----------------------------
    @action(detail=True, methods=["get"], url_path="ficha/pdf")
    def ficha_pdf(self, request, pk=None):
        if not _is_staff(request.user):
            return Response({"detail": "Solo staff."}, status=403)

        it = self.get_object()

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            title=f"Ficha {it.codigo}",
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Ficha de inventario: <b>{it.codigo}</b>", styles["Title"]),
            Spacer(1, 10),
        ]

        # Foto grande
        if it.foto and hasattr(it.foto, "path"):
            try:
                story.append(RLImage(it.foto.path, width=220, height=220))
                story.append(Spacer(1, 10))
            except Exception:
                pass

        data = [
            ["Categoría", it.categoria.nombre if it.categoria_id else ""],
            ["Ubicación", it.ubicacion.nombre if it.ubicacion_id else ""],
            ["Estado", it.get_estado_display()],
            ["Marca", it.marca or ""],
            ["Modelo", it.modelo or ""],
            ["Serie", it.serie or ""],
            ["Etiqueta interna", it.etiqueta_interna or ""],
            ["Responsable", it.responsable or ""],
            ["Activo", "Sí" if it.activo else "No"],
            ["Precio sugerido venta", f"${it.precio_sugerido_venta:.2f}" if it.precio_sugerido_venta is not None else ""],
            ["Fecha alta", it.fecha_alta.isoformat() if it.fecha_alta else ""],
            ["Fecha baja", it.fecha_baja.isoformat() if it.fecha_baja else ""],
            ["Motivo baja", it.motivo_baja.nombre if it.motivo_baja_id else ""],
        ]

        table = Table(data, colWidths=[160, 340])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                ]
            )
        )

        story.append(table)
        doc.build(story)

        pdf = buffer.getvalue()
        buffer.close()

        filename = f"ficha_{it.codigo}_{localdate().isoformat()}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
